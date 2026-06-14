from flask import Flask, jsonify, request, render_template, Response
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from threading import Lock
import os, sys, secrets, json, urllib.request, urllib.error, urllib.parse, io, re, base64

APP_VERSION = '0.1.27'
GITHUB_REPO = 'federicoroldos/clarifi'

# Models used to structure raw OCR text into transaction fields when the user has
# saved an AI key. Only the OCR *text* is sent — never the image. The provider is
# auto-detected from the key prefix: Groq keys start with 'gsk_', Anthropic keys
# with 'sk-ant-', everything else is treated as a Google Gemini key.
GEMINI_STRUCTURE_MODEL = 'gemini-2.0-flash'
GROQ_STRUCTURE_MODEL = 'llama-3.3-70b-versatile'
CLAUDE_STRUCTURE_MODEL = 'claude-haiku-4-5-20251001'
# Vision models. Gemini 2.0 Flash and Claude Haiku 4.5 are already multimodal, so
# they read images with the same model id above. Groq's text model is text-only, so
# image input uses a separate vision-capable model.
GROQ_VISION_MODEL = 'meta-llama/llama-4-scout-17b-16e-instruct'
VISION_MAX_DIM = 1600    # downscale an image's longest side before sending to vision
VISION_MAX_PAGES = 8     # scanned-statement pages sent to a vision model at most
# Groq and Anthropic sit behind Cloudflare, which 403s the default
# 'Python-urllib/x' User-Agent as a suspected bot. Send a real UA on every AI call.
AI_USER_AGENT = 'ClariFi/' + APP_VERSION


def _default_data_path():
    if os.environ.get('DATA_PATH'):
        return os.environ['DATA_PATH']
    if getattr(sys, 'frozen', False):
        base = os.path.join(os.environ.get('APPDATA') or os.path.expanduser('~'), 'ClariFi')
        os.makedirs(base, exist_ok=True)
        return os.path.join(base, 'finance_data.xlsx')
    return 'finance_data.xlsx'

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['TEMPLATES_AUTO_RELOAD'] = True
DATA_PATH  = _default_data_path()
XLSX_LOCK  = Lock()
CATEGORIES = ['Supermarket','Food','Transport','Games','Services','Health','Hanging out','Others']
CURRENCIES = {
    'krw': {'name': 'Korean Won',    'symbol': '₩',   'decimals': 0},
    'uyu': {'code': 'UYU', 'name': 'Uruguayan Peso',  'symbol': '$U',  'decimals': 2},
    'usd': {'code': 'USD', 'name': 'US Dollar',       'symbol': 'US$', 'decimals': 2},
    'eur': {'code': 'EUR', 'name': 'Euro',            'symbol': '€',   'decimals': 2},
    'ars': {'code': 'ARS', 'name': 'Argentine Peso',  'symbol': 'AR$', 'decimals': 2},
}
CURRENCIES['krw'].update({'code': 'KRW', 'symbol': '₩'})
LEGACY_ACCOUNT_BANKS = {
    'krw': 'Korean Won Account',
    'uyu': 'Uruguayan Peso Account',
    'usd': 'US Dollar Account',
}
LEGACY_CURRENCIES = ('krw', 'uyu', 'usd')
DEFAULT_NEW_ACCOUNTS = (
    ('usd', 'US Dollar Account'),
    ('eur', 'Euro Account'),
)

# XLSX storage
SHEETS = {
    'config': ['key', 'value'],
    'accounts': ['id', 'bank', 'currency', 'balance', 'created_at', 'archived', 'color'],
    'transactions': ['id', 'date', 'description', 'amount', 'category', 'type', 'account', 'transfer_id', 'counterpart', 'transfer_dir'],
    'fixed_payments': ['id', 'name', 'amount', 'account', 'category', 'day', 'type'],
    'fixed_applied': ['payment_id', 'year_month'],
}

def _headers(ws):
    return [c.value for c in ws[1]]

def _ensure_headers(ws, expected):
    if _headers(ws)[:len(expected)] != expected:
        for idx, name in enumerate(expected, start=1):
            ws.cell(row=1, column=idx, value=name)

def _rows(ws):
    headers = _headers(ws)
    out = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in values):
            continue
        out.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    return out

def _next_id(ws):
    ids = []
    for row in _rows(ws):
        try:
            ids.append(int(row.get('id') or 0))
        except (TypeError, ValueError):
            pass
    return max(ids, default=0) + 1

def _load_wb():
    return load_workbook(DATA_PATH)


# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html', app_version=APP_VERSION)


@app.route('/favicon.ico')
def favicon():
    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64">'
        '<rect width="64" height="64" rx="18" fill="#1c1c20"/>'
        '<circle cx="32" cy="32" r="20" fill="none" stroke="#10b981" stroke-width="2.5" opacity="0.32"/>'
        '<circle cx="32" cy="32" r="14" fill="none" stroke="#10b981" stroke-width="2.5" opacity="0.6"/>'
        '<circle cx="32" cy="32" r="7" fill="#10b981"/>'
        '<circle cx="32" cy="32" r="2.4" fill="#08080a"/>'
        '</svg>'
    )
    return Response(svg, mimetype='image/svg+xml')

@app.route('/api/summary')
def api_summary(): return jsonify(build_summary())

@app.route('/api/transactions')
def api_transactions():
    with XLSX_LOCK:
        wb = _load_wb()
        accounts = _accounts_from_wb(wb, include_archived=True)
        txns = _rows(wb['transactions'])
    accounts_map = {account['id']: account for account in accounts}
    annotated = [t for t in (_annotate_txn(raw, accounts_map) for raw in txns) if t]
    annotated.sort(key=lambda t: (str(t.get('date') or ''), int(t.get('id') or 0)), reverse=True)
    return jsonify(annotated)

@app.route('/api/fund',    methods=['POST'])
def add_fund():    return _add_txn(request.json, 'fund')

@app.route('/api/expense', methods=['POST'])
def add_expense(): return _add_txn(request.json, 'expense')





# fixed payments


@app.route('/api/fixed/<int:fid>', methods=['DELETE'])
def delete_fixed(fid):
    with XLSX_LOCK:
        wb = _load_wb()
        fixed_ws = wb['fixed_payments']
        for row_idx in range(fixed_ws.max_row, 1, -1):
            if int(fixed_ws.cell(row=row_idx, column=1).value or 0) == fid:
                fixed_ws.delete_rows(row_idx, 1)

        applied_ws = wb['fixed_applied']
        for row_idx in range(applied_ws.max_row, 1, -1):
            if int(applied_ws.cell(row=row_idx, column=1).value or 0) == fid:
                applied_ws.delete_rows(row_idx, 1)
        wb.save(DATA_PATH)
    return jsonify({'ok': True})

@app.route('/api/fixed/<int:fid>/apply', methods=['POST'])
def apply_fixed(fid):
    this_month = datetime.now().strftime('%Y-%m')
    today_str  = datetime.now().strftime('%Y-%m-%d')
    with XLSX_LOCK:
        wb = _load_wb()
        fixed = _rows(wb['fixed_payments'])
        fp = next((row for row in fixed if int(row.get('id') or 0) == fid), None)
        if not fp:
            return jsonify({'ok': False}), 404

        applied_ws = wb['fixed_applied']
        applied = _rows(applied_ws)
        if any(int(a.get('payment_id') or 0) == fid and a.get('year_month') == this_month for a in applied):
            return jsonify({'ok': False, 'error': 'already applied this month'}), 400
        applied_ws.append([fid, this_month])
        wb.save(DATA_PATH)
    return _add_txn({'amount': fp['amount'], 'description': fp['name'],
                     'account': fp['account'], 'category': fp['category'],
                     'date': today_str}, _fixed_type(fp.get('type')))

@app.route('/api/fixed/<int:fid>/undo', methods=['POST'])
def undo_fixed(fid):
    this_month = datetime.now().strftime('%Y-%m')
    with XLSX_LOCK:
        wb = _load_wb()
        fixed = _rows(wb['fixed_payments'])
        fp = next((row for row in fixed if int(row.get('id') or 0) == fid), None)
        if not fp:
            return jsonify({'ok': False}), 404

        ftype = _fixed_type(fp.get('type'))
        txn_ws = wb['transactions']
        txns = _rows(txn_ws)
        matches = [
            t for t in txns
            if t.get('description') == fp.get('name')
            and t.get('account') == fp.get('account')
            and t.get('type') == ftype
            and str(t.get('date') or '').startswith(this_month)
        ]
        t = max(matches, key=lambda row: int(row.get('id') or 0), default=None)
        if t:
            for row_idx in range(txn_ws.max_row, 1, -1):
                if int(txn_ws.cell(row=row_idx, column=1).value or 0) == int(t.get('id') or 0):
                    txn_ws.delete_rows(row_idx, 1)
                    break

        applied_ws = wb['fixed_applied']
        for row_idx in range(applied_ws.max_row, 1, -1):
            if int(applied_ws.cell(row=row_idx, column=1).value or 0) == fid and applied_ws.cell(row=row_idx, column=2).value == this_month:
                applied_ws.delete_rows(row_idx, 1)
        wb.save(DATA_PATH)

    if t:
        acc = t.get('account')
        bal = get_balances()[acc]
        amount = float(t.get('amount') or 0)
        delta = -amount if t.get('type') == 'fund' else amount
        set_balance(acc, round_acc(acc, bal + delta))
    return jsonify({'ok': True})

# Modern account model overrides
CURRENCIES['krw']['symbol'] = '\u20a9'
for _currency_key, _currency_meta in CURRENCIES.items():
    _currency_meta.setdefault('code', _currency_key.upper())

def round_currency(currency, val):
    currency = str(currency or '').strip().lower()
    if currency not in CURRENCIES:
        raise ValueError('unknown currency')
    return round(float(val), CURRENCIES[currency]['decimals'])

def init_data():
    if os.path.exists(DATA_PATH):
        wb = load_workbook(DATA_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for sheet, headers in SHEETS.items():
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        _ensure_headers(ws, headers)

    config = wb['config']
    existing = {r.get('key') for r in _rows(config)}
    for currency in LEGACY_CURRENCIES:
        key = f'balance_{currency}'
        if key not in existing:
            config.append([key, 0])

    accounts_ws = wb['accounts']
    if not _rows(accounts_ws):
        legacy_balances = {
            str(r.get('key') or '').replace('balance_', ''): float(r.get('value') or 0)
            for r in _rows(config)
            if str(r.get('key') or '').startswith('balance_')
        }
        if any(v > 0 for v in legacy_balances.values()):
            # Migrating an old single-currency install: keep KRW/UYU/USD accounts.
            for currency in LEGACY_CURRENCIES:
                accounts_ws.append([
                    currency,
                    LEGACY_ACCOUNT_BANKS[currency],
                    currency,
                    round_currency(currency, legacy_balances.get(currency, 0)),
                    datetime.now().isoformat(timespec='seconds'),
                    False,
                    DEFAULT_ACC_COLORS.get(currency, '#4a90f8'),
                ])
        else:
            # Fresh install: seed only USD and EUR.
            for currency, bank in DEFAULT_NEW_ACCOUNTS:
                accounts_ws.append([
                    currency,
                    bank,
                    currency,
                    0,
                    datetime.now().isoformat(timespec='seconds'),
                    False,
                    DEFAULT_ACC_COLORS.get(currency, '#4a90f8'),
                ])

    wb.save(DATA_PATH)

def _is_archived(value):
    return str(value).lower() in ('1', 'true', 'yes')

def _fixed_type(value):
    t = str(value or '').strip().lower()
    return t if t in ('fund', 'expense') else 'expense'

def _currency_id(value):
    currency = str(value or '').strip().lower()
    if currency not in CURRENCIES:
        raise ValueError('unknown currency')
    return currency

DEFAULT_ACC_COLORS = {
    'uyu': '#4a90f8',
    'usd': '#32d74b',
    'krw': '#bf5af2',
    'eur': '#5ac8fa',
    'ars': '#ff9f0a',
}

def _account_json(row):
    currency = _currency_id(row.get('currency') or 'uyu')
    color = str(row.get('color') or '').strip() or DEFAULT_ACC_COLORS.get(currency, '#4a90f8')
    return {
        'id': str(row.get('id') or ''),
        'bank': str(row.get('bank') or '').strip() or 'Account',
        'currency': currency,
        'balance': round_currency(currency, row.get('balance') or 0),
        'created_at': row.get('created_at') or '',
        'archived': _is_archived(row.get('archived')),
        'currency_meta': CURRENCIES[currency],
        'color': color,
    }

def _accounts_from_wb(wb, include_archived=False):
    accounts = []
    for row in _rows(wb['accounts']):
        try:
            account = _account_json(row)
        except ValueError:
            continue
        if account['id'] and (include_archived or not account['archived']):
            accounts.append(account)
    return accounts

def get_accounts(include_archived=False):
    with XLSX_LOCK:
        wb = _load_wb()
        return _accounts_from_wb(wb, include_archived)

def get_accounts_map(include_archived=False):
    return {account['id']: account for account in get_accounts(include_archived)}

def get_account(account_id):
    return get_accounts_map().get(str(account_id or ''))

def _default_account_id():
    accounts = get_accounts()
    if not accounts:
        raise ValueError('no accounts configured')
    return 'uyu' if any(account['id'] == 'uyu' for account in accounts) else accounts[0]['id']

def _new_account_id(existing_ids):
    while True:
        account_id = 'acct_' + secrets.token_hex(4)
        if account_id not in existing_ids:
            return account_id

def get_balances():
    return {account['id']: account['balance'] for account in get_accounts()}

def set_balance(account_id, val):
    account_id = str(account_id or '')
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['accounts']
        headers = _headers(ws)
        id_col = headers.index('id') + 1
        currency_col = headers.index('currency') + 1
        balance_col = headers.index('balance') + 1
        rounded = None
        for row_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row=row_idx, column=id_col).value or '') == account_id:
                currency = ws.cell(row=row_idx, column=currency_col).value
                rounded = round_currency(currency, val)
                ws.cell(row=row_idx, column=balance_col, value=rounded)
                break
        else:
            raise ValueError('unknown account')

        if account_id in CURRENCIES:
            key = f'balance_{account_id}'
            config = wb['config']
            for row_idx in range(2, config.max_row + 1):
                if config.cell(row=row_idx, column=1).value == key:
                    config.cell(row=row_idx, column=2, value=rounded)
                    break
            else:
                config.append([key, rounded])
        wb.save(DATA_PATH)
    return rounded

def round_acc(account_id, val):
    account = get_account(account_id)
    if not account:
        raise ValueError('unknown account')
    return round_currency(account['currency'], val)

def _blank_stats():
    return {'exp_cat': {}, 'monthly': {}, 'last30': 0.0, 'in30': 0.0, 'total_txns': 0}

def _annotate_txn(txn, accounts):
    account = accounts.get(str(txn.get('account') or ''))
    if not account:
        return None
    item = dict(txn)
    item['account_name'] = account['bank']
    item['currency'] = account['currency']
    counterpart_id = str(txn.get('counterpart') or '')
    if counterpart_id:
        counterpart = accounts.get(counterpart_id)
        if counterpart:
            item['counterpart_name'] = counterpart['bank']
            item['counterpart_currency'] = counterpart['currency']
    return item

def build_summary():
    with XLSX_LOCK:
        wb = _load_wb()
        accounts = _accounts_from_wb(wb)
        all_accounts = _accounts_from_wb(wb, include_archived=True)
        txns = _rows(wb['transactions'])
        fixed = _rows(wb['fixed_payments'])
        applied = _rows(wb['fixed_applied'])

    accounts_map = {account['id']: account for account in accounts}
    txns.sort(key=lambda t: (str(t.get('date') or ''), int(t.get('id') or 0)), reverse=True)
    fixed.sort(key=lambda f: int(f.get('day') or 0))

    today_str = datetime.now().strftime('%Y-%m-%d')
    this_month = today_str[:7]
    today_day = int(today_str[8:])
    cutoff = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    applied_set = {(a['payment_id'], a['year_month']) for a in applied}

    stats = {account['id']: _blank_stats() for account in accounts}
    overview = {
        'by_currency': {
            currency: {'balance': 0.0, 'last30': 0.0, 'in30': 0.0, 'accounts': 0}
            for currency in CURRENCIES
        },
        'monthly': {},
        'exp_cat': {},
        'total_txns': len(txns),
    }

    for account in accounts:
        bucket = overview['by_currency'][account['currency']]
        bucket['balance'] += account['balance']
        bucket['accounts'] += 1

    recent = []
    for raw_txn in txns:
        txn = _annotate_txn(raw_txn, accounts_map)
        if not txn:
            continue
        recent.append(txn)
        account_id = txn['account']
        currency = txn['currency']
        amount = float(txn.get('amount') or 0)
        date_str = str(txn.get('date') or '')
        month = date_str[:7]
        stats[account_id]['total_txns'] += 1
        ttype = txn.get('type')
        if ttype == 'transfer':
            continue
        if ttype == 'expense':
            category = txn.get('category') or 'Others'
            stats[account_id]['exp_cat'][category] = stats[account_id]['exp_cat'].get(category, 0) + amount
            overview['exp_cat'][category] = overview['exp_cat'].get(category, 0) + amount
            if date_str >= cutoff:
                stats[account_id]['last30'] += amount
                overview['by_currency'][currency]['last30'] += amount
        elif date_str >= cutoff:
            stats[account_id]['in30'] += amount
            overview['by_currency'][currency]['in30'] += amount
        if month:
            stats[account_id]['monthly'].setdefault(month, {'in': 0.0, 'out': 0.0})
            overview['monthly'].setdefault(month, {})
            overview['monthly'][month].setdefault(currency, {'in': 0.0, 'out': 0.0})
            direction = 'in' if ttype == 'fund' else 'out'
            stats[account_id]['monthly'][month][direction] += amount
            overview['monthly'][month][currency][direction] += amount

    for account_id in stats:
        months = sorted(stats[account_id]['monthly'])[-6:]
        stats[account_id]['monthly'] = {month: stats[account_id]['monthly'][month] for month in months}
    months = sorted(overview['monthly'])[-6:]
    overview['monthly'] = {month: overview['monthly'][month] for month in months}

    normalized_fixed = []
    for fixed_payment in fixed:
        account = accounts_map.get(str(fixed_payment.get('account') or ''))
        if not account:
            continue
        item = dict(fixed_payment)
        item['account_name'] = account['bank']
        item['currency'] = account['currency']
        item['type'] = _fixed_type(fixed_payment.get('type'))
        item['applied_this_month'] = (item['id'], this_month) in applied_set
        item['due_this_month'] = item['day'] <= today_day and not item['applied_this_month']
        normalized_fixed.append(item)

    return {
        'balances': get_balances(),
        'stats': stats,
        'overview': overview,
        'recent': recent[:15],
        'fixed': normalized_fixed,
        'due_count': sum(1 for item in normalized_fixed if item['due_this_month']),
        'categories': CATEGORIES,
        'accounts': accounts_map,
        'account_list': accounts,
        'all_account_list': all_accounts,
        'currencies': CURRENCIES,
        'total_txns': len(txns),
    }


def modern_accounts():
    include_archived = str(request.args.get('include_archived') or '').lower() in ('1', 'true', 'yes')
    return jsonify(get_accounts(include_archived=include_archived))

def modern_create_account():
    data = request.json or {}
    bank = str(data.get('bank') or '').strip()
    if not bank:
        return jsonify({'ok': False, 'error': 'bank is required'}), 400
    try:
        currency = _currency_id(data.get('currency'))
        balance = round_currency(currency, data.get('balance') or 0)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'invalid currency or balance'}), 400

    color = str(data.get('color') or '').strip() or DEFAULT_ACC_COLORS.get(currency, '#4a90f8')

    with XLSX_LOCK:
        wb = _load_wb()
        existing_ids = {account['id'] for account in _accounts_from_wb(wb, include_archived=True)}
        account_id = _new_account_id(existing_ids)
        wb['accounts'].append([
            account_id,
            bank,
            currency,
            balance,
            datetime.now().isoformat(timespec='seconds'),
            False,
            color,
        ])
        wb.save(DATA_PATH)

    return jsonify({'ok': True, 'account': get_account(account_id)})


def modern_edit_account(account_id):
    account_id = str(account_id or '')
    data = request.json or {}
    bank = str(data.get('bank') or '').strip()
    if not bank:
        return jsonify({'ok': False, 'error': 'bank name is required'}), 400

    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['accounts']
        headers = _headers(ws)
        id_col = headers.index('id') + 1
        found = None
        for row_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row=row_idx, column=id_col).value or '') == account_id:
                found = row_idx
                break
        if found is None:
            return jsonify({'ok': False, 'error': 'account not found'}), 404

        ws.cell(row=found, column=headers.index('bank') + 1, value=bank)

        new_currency = data.get('currency')
        if new_currency is not None:
            try:
                currency = _currency_id(new_currency)
                ws.cell(row=found, column=headers.index('currency') + 1, value=currency)
            except (ValueError, KeyError):
                return jsonify({'ok': False, 'error': 'invalid currency'}), 400
        else:
            currency = str(ws.cell(row=found, column=headers.index('currency') + 1).value or 'uyu')

        if data.get('balance') is not None:
            try:
                rounded = round_currency(currency, data['balance'])
                ws.cell(row=found, column=headers.index('balance') + 1, value=rounded)
            except (ValueError, TypeError):
                return jsonify({'ok': False, 'error': 'invalid balance'}), 400

        new_color = str(data.get('color') or '').strip()
        if new_color and 'color' in headers:
            ws.cell(row=found, column=headers.index('color') + 1, value=new_color)

        wb.save(DATA_PATH)
    return jsonify({'ok': True, 'account': get_account(account_id)})


def modern_delete_account(account_id):
    account_id = str(account_id or '')
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['accounts']
        headers = _headers(ws)
        id_col = headers.index('id') + 1
        archived_col = headers.index('archived') + 1
        found = None
        for row_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row=row_idx, column=id_col).value or '') == account_id:
                found = row_idx
                break
        if found is None:
            return jsonify({'ok': False, 'error': 'account not found'}), 404
        if _is_archived(ws.cell(row=found, column=archived_col).value):
            return jsonify({'ok': True})

        ws.cell(row=found, column=archived_col, value=True)
        wb.save(DATA_PATH)

    return jsonify({'ok': True})

def modern_permanent_delete_account(account_id):
    account_id = str(account_id or '')
    with XLSX_LOCK:
        wb = _load_wb()
        accounts_ws = wb['accounts']
        headers = _headers(accounts_ws)
        id_col = headers.index('id') + 1
        archived_col = headers.index('archived') + 1
        found = None
        for row_idx in range(2, accounts_ws.max_row + 1):
            if str(accounts_ws.cell(row=row_idx, column=id_col).value or '') == account_id:
                found = row_idx
                break
        if found is None:
            return jsonify({'ok': False, 'error': 'account not found'}), 404
        if not _is_archived(accounts_ws.cell(row=found, column=archived_col).value):
            return jsonify({'ok': False, 'error': 'deactivate account before deleting it'}), 400

        fixed_ids = set()
        fixed_ws = wb['fixed_payments']
        fixed_headers = _headers(fixed_ws)
        fixed_id_col = fixed_headers.index('id') + 1
        fixed_account_col = fixed_headers.index('account') + 1
        for row_idx in range(fixed_ws.max_row, 1, -1):
            if str(fixed_ws.cell(row=row_idx, column=fixed_account_col).value or '') == account_id:
                fixed_ids.add(int(fixed_ws.cell(row=row_idx, column=fixed_id_col).value or 0))
                fixed_ws.delete_rows(row_idx, 1)

        applied_ws = wb['fixed_applied']
        for row_idx in range(applied_ws.max_row, 1, -1):
            if int(applied_ws.cell(row=row_idx, column=1).value or 0) in fixed_ids:
                applied_ws.delete_rows(row_idx, 1)

        txn_ws = wb['transactions']
        txn_headers = _headers(txn_ws)
        txn_account_col = txn_headers.index('account') + 1
        for row_idx in range(txn_ws.max_row, 1, -1):
            if str(txn_ws.cell(row=row_idx, column=txn_account_col).value or '') == account_id:
                txn_ws.delete_rows(row_idx, 1)

        if account_id in CURRENCIES:
            config_ws = wb['config']
            key = f'balance_{account_id}'
            for row_idx in range(config_ws.max_row, 1, -1):
                if config_ws.cell(row=row_idx, column=1).value == key:
                    config_ws.delete_rows(row_idx, 1)

        accounts_ws.delete_rows(found, 1)
        wb.save(DATA_PATH)

    return jsonify({'ok': True})

def _add_txn(data, txn_type):
    data = data or {}
    account_id = str(data.get('account') or _default_account_id())
    if not get_account(account_id):
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    try:
        amount = round_acc(account_id, abs(float(data.get('amount') or 0)))
    except ValueError:
        return jsonify({'ok': False, 'error': 'invalid amount'}), 400
    if amount <= 0:
        return jsonify({'ok': False, 'error': 'amount must be greater than zero'}), 400

    balances = get_balances()
    new_balance = balances[account_id] + amount if txn_type == 'fund' else balances[account_id] - amount
    new_balance = set_balance(account_id, new_balance)
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        ws.append([
            _next_id(ws),
            data.get('date', datetime.now().strftime('%Y-%m-%d')),
            data.get('description', ''),
            amount,
            data.get('category', 'Others'),
            txn_type,
            account_id,
        ])
        wb.save(DATA_PATH)
    return jsonify({'ok': True, 'balance': new_balance})

def modern_delete_txn(tid):
    legs = []
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        headers = _headers(ws)
        found = None
        for row_idx in range(2, ws.max_row + 1):
            if int(ws.cell(row=row_idx, column=1).value or 0) == tid:
                found = row_idx
                break
        if found is None:
            return jsonify({'ok': False}), 404
        primary = {headers[col - 1]: ws.cell(row=found, column=col).value for col in range(1, len(headers) + 1)}

        transfer_id = primary.get('transfer_id') if primary.get('type') == 'transfer' else None
        if transfer_id and 'transfer_id' in headers:
            tid_col = headers.index('transfer_id') + 1
            for row_idx in range(ws.max_row, 1, -1):
                if str(ws.cell(row=row_idx, column=tid_col).value or '') == str(transfer_id):
                    legs.append({headers[col - 1]: ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)})
                    ws.delete_rows(row_idx, 1)
        else:
            legs.append(primary)
            ws.delete_rows(found, 1)
        wb.save(DATA_PATH)

    for leg in legs:
        account_id = str(leg.get('account') or '')
        if not get_account(account_id):
            continue
        amount = float(leg.get('amount') or 0)
        balance = get_balances()[account_id]
        ltype = leg.get('type')
        if ltype == 'transfer':
            direction = str(leg.get('transfer_dir') or '').lower()
            new_balance = balance + amount if direction == 'out' else balance - amount
        elif ltype == 'fund':
            new_balance = balance - amount
        else:
            new_balance = balance + amount
        set_balance(account_id, new_balance)
    return jsonify({'ok': True})

def modern_edit_txn(tid):
    data = request.json or {}
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        headers = _headers(ws)
        col = {h: i + 1 for i, h in enumerate(headers)}
        found = None
        for row_idx in range(2, ws.max_row + 1):
            if int(ws.cell(row=row_idx, column=col['id']).value or 0) == tid:
                found = row_idx
                break
        if found is None:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        old = {h: ws.cell(row=found, column=col[h]).value for h in headers}

    if old.get('type') == 'transfer':
        return jsonify({'ok': False, 'error': 'transfers cannot be edited; delete and recreate'}), 400

    old_account = str(old.get('account') or '')
    old_amount = float(old.get('amount') or 0)
    old_type = old.get('type')

    new_account = str(data.get('account') or old_account)
    if not get_account(new_account):
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    try:
        new_amount = round_acc(new_account, abs(float(data.get('amount') or 0)))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'invalid amount'}), 400
    if new_amount <= 0:
        return jsonify({'ok': False, 'error': 'amount must be greater than zero'}), 400

    if get_account(old_account):
        bal = get_balances()[old_account]
        bal = bal - old_amount if old_type == 'fund' else bal + old_amount
        set_balance(old_account, bal)

    bal = get_balances()[new_account]
    bal = bal + new_amount if old_type == 'fund' else bal - new_amount
    set_balance(new_account, bal)

    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        headers = _headers(ws)
        col = {h: i + 1 for i, h in enumerate(headers)}
        for row_idx in range(2, ws.max_row + 1):
            if int(ws.cell(row=row_idx, column=col['id']).value or 0) == tid:
                ws.cell(row=row_idx, column=col['date']).value = data.get('date') or old.get('date')
                ws.cell(row=row_idx, column=col['description']).value = data.get('description', '')
                ws.cell(row=row_idx, column=col['amount']).value = new_amount
                ws.cell(row=row_idx, column=col['category']).value = data.get('category', 'Others')
                ws.cell(row=row_idx, column=col['account']).value = new_account
                break
        wb.save(DATA_PATH)
    return jsonify({'ok': True})

def modern_set_balance():
    data = request.json or {}
    account_id = str(data.get('account') or '')
    if not get_account(account_id):
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    balance = set_balance(account_id, float(data.get('balance') or 0))
    return jsonify({'ok': True, 'balance': balance})

def modern_fixed():
    with XLSX_LOCK:
        wb = _load_wb()
        rows = _rows(wb['fixed_payments'])
        applied = _rows(wb['fixed_applied'])
    accounts = get_accounts_map()
    this_month = datetime.now().strftime('%Y-%m')
    today_day = datetime.now().day
    applied_set = {(a['payment_id'], a['year_month']) for a in applied}
    result = []
    for row in sorted(rows, key=lambda item: int(item.get('day') or 0)):
        account = accounts.get(str(row.get('account') or ''))
        if not account:
            continue
        item = dict(row)
        item['account_name'] = account['bank']
        item['currency'] = account['currency']
        item['type'] = _fixed_type(row.get('type'))
        item['applied_this_month'] = (item['id'], this_month) in applied_set
        item['due_this_month'] = item['day'] <= today_day and not item['applied_this_month']
        result.append(item)
    return jsonify(result)

def modern_create_fixed():
    data = request.json or {}
    account_id = str(data.get('account') or _default_account_id())
    if not get_account(account_id):
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    day = int(data.get('day', 1))
    if not 1 <= day <= 31:
        return jsonify({'ok': False, 'error': 'day must be 1-31'}), 400
    amount = round_acc(account_id, float(data.get('amount', 0)))
    ftype = _fixed_type(data.get('type'))
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['fixed_payments']
        fixed_id = _next_id(ws)
        ws.append([fixed_id, data.get('name', ''), amount, account_id, data.get('category', 'Others'), day, ftype])
        wb.save(DATA_PATH)
    return jsonify({'ok': True, 'id': fixed_id})

def modern_edit_fixed(fid):
    data = request.json or {}
    account_id = str(data.get('account') or _default_account_id())
    if not get_account(account_id):
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    try:
        day = int(data.get('day', 1))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'invalid day'}), 400
    if not 1 <= day <= 31:
        return jsonify({'ok': False, 'error': 'day must be 1-31'}), 400
    try:
        amount = round_acc(account_id, float(data.get('amount', 0)))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'invalid amount'}), 400
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['fixed_payments']
        headers = _headers(ws)
        col = {h: headers.index(h) + 1 for h in headers}
        found = False
        for row_idx in range(2, ws.max_row + 1):
            if int(ws.cell(row=row_idx, column=col['id']).value or 0) == fid:
                ws.cell(row=row_idx, column=col['name']).value = data.get('name', '')
                ws.cell(row=row_idx, column=col['amount']).value = amount
                ws.cell(row=row_idx, column=col['account']).value = account_id
                ws.cell(row=row_idx, column=col['category']).value = data.get('category', 'Others')
                ws.cell(row=row_idx, column=col['day']).value = day
                if 'type' in col:
                    ws.cell(row=row_idx, column=col['type']).value = _fixed_type(data.get('type'))
                found = True
                break
        if not found:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        wb.save(DATA_PATH)
    return jsonify({'ok': True})

def modern_export():
    with XLSX_LOCK:
        wb = _load_wb()
        config = {r.get('key'): r.get('value') for r in _rows(wb['config']) if r.get('key') != 'ai_api_key'}
        accounts = _accounts_from_wb(wb, include_archived=True)
        txns = _rows(wb['transactions'])
        fixed = _rows(wb['fixed_payments'])
        applied_rows = _rows(wb['fixed_applied'])

    applied = {}
    for row in applied_rows:
        month = row.get('year_month')
        payment_id = row.get('payment_id')
        if month:
            applied.setdefault(month, []).append(payment_id)

    payload = {
        'version': 2,
        'exported_at': datetime.now().isoformat(),
        'accounts': accounts,
        'config': config,
        'txns': txns,
        'fixed': fixed,
        'applied': applied,
    }

    home = os.path.expanduser('~')
    desktop = os.path.join(home, 'Desktop')
    target_dir = desktop if os.path.isdir(desktop) else home
    filename = 'clarifi-' + datetime.now().strftime('%Y-%m-%d') + '.json'
    full_path = os.path.join(target_dir, filename)
    try:
        with open(full_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
    except OSError as e:
        return jsonify({'ok': False, 'error': 'could not write file: ' + str(e)}), 500

    return jsonify({'ok': True, 'path': full_path, 'filename': filename})

def modern_import():
    data = request.json or {}
    accounts = data.get('accounts') if isinstance(data.get('accounts'), list) else []
    config = data.get('config') or {}
    txns = data.get('txns') if isinstance(data.get('txns'), list) else []
    fixed = data.get('fixed') if isinstance(data.get('fixed'), list) else []
    applied = data.get('applied') if isinstance(data.get('applied'), dict) else {}

    with XLSX_LOCK:
        wb = _load_wb()
        for sheet, headers in SHEETS.items():
            ws = wb[sheet]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            _ensure_headers(ws, headers)

        imported_ids = []
        if accounts:
            seen = set()
            for row in accounts:
                try:
                    account_id = str(row.get('id') or '').strip()
                    currency = _currency_id(row.get('currency'))
                    if not account_id or account_id in seen:
                        continue
                    balance = round_currency(currency, row.get('balance') or 0)
                except (TypeError, ValueError):
                    continue
                seen.add(account_id)
                imported_ids.append(account_id)
                color = str(row.get('color') or '').strip() or DEFAULT_ACC_COLORS.get(currency, '#4a90f8')
                wb['accounts'].append([
                    account_id,
                    str(row.get('bank') or '').strip() or 'Account',
                    currency,
                    balance,
                    row.get('created_at') or datetime.now().isoformat(timespec='seconds'),
                    _is_archived(row.get('archived')),
                    color,
                ])
        else:
            for currency in LEGACY_CURRENCIES:
                imported_ids.append(currency)
                wb['accounts'].append([
                    currency,
                    LEGACY_ACCOUNT_BANKS[currency],
                    currency,
                    round_currency(currency, config.get(f'balance_{currency}') or 0),
                    datetime.now().isoformat(timespec='seconds'),
                    False,
                    DEFAULT_ACC_COLORS.get(currency, '#4a90f8'),
                ])

        if not imported_ids:
            return jsonify({'ok': False, 'error': 'no valid accounts'}), 400

        accounts_map = {row['id']: row for row in _accounts_from_wb(wb, include_archived=True)}
        for currency in LEGACY_CURRENCIES:
            account = accounts_map.get(currency)
            wb['config'].append([f'balance_{currency}', account['balance'] if account else 0])

        txns_ws = wb['transactions']
        for row in txns:
            account_id = str(row.get('account') or '')
            if account_id not in imported_ids:
                account_id = imported_ids[0]
            txn_type = row.get('type') if row.get('type') in ('fund', 'expense', 'transfer') else 'expense'
            counterpart = str(row.get('counterpart') or '') or None
            if counterpart and counterpart not in imported_ids:
                counterpart = None
            transfer_dir = str(row.get('transfer_dir') or '').lower()
            if transfer_dir not in ('in', 'out'):
                transfer_dir = None
            txns_ws.append([
                int(row.get('id') or _next_id(txns_ws)),
                row.get('date') or datetime.now().strftime('%Y-%m-%d'),
                row.get('description') or '',
                round_currency(accounts_map[account_id]['currency'], float(row.get('amount') or 0)),
                row.get('category') or 'Others',
                txn_type,
                account_id,
                str(row.get('transfer_id') or '') or None,
                counterpart,
                transfer_dir,
            ])

        fixed_ws = wb['fixed_payments']
        for row in fixed:
            account_id = str(row.get('account') or '')
            if account_id not in imported_ids:
                account_id = imported_ids[0]
            fixed_ws.append([
                int(row.get('id') or _next_id(fixed_ws)),
                row.get('name') or '',
                round_currency(accounts_map[account_id]['currency'], float(row.get('amount') or 0)),
                account_id,
                row.get('category') or 'Others',
                min(31, max(1, int(row.get('day') or 1))),
                _fixed_type(row.get('type')),
            ])

        applied_ws = wb['fixed_applied']
        for month, ids in applied.items():
            if isinstance(ids, list):
                for payment_id in ids:
                    applied_ws.append([int(payment_id), month])

        wb.save(DATA_PATH)

    return jsonify({'ok': True})

def modern_clear():
    with XLSX_LOCK:
        wb = _load_wb()
        for sheet, headers in SHEETS.items():
            ws = wb[sheet]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            _ensure_headers(ws, headers)

        config = wb['config']
        for currency in LEGACY_CURRENCIES:
            config.append([f'balance_{currency}', 0])

        accounts_ws = wb['accounts']
        for currency, bank in DEFAULT_NEW_ACCOUNTS:
            accounts_ws.append([
                currency,
                bank,
                currency,
                0,
                datetime.now().isoformat(timespec='seconds'),
                False,
                DEFAULT_ACC_COLORS.get(currency, '#4a90f8'),
            ])

        wb.save(DATA_PATH)
    return jsonify({'ok': True})

def _save_fxrate(from_cur, to_cur, rate):
    key = f'fxrate_{from_cur}_{to_cur}'
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['config']
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == key:
                ws.cell(row=row_idx, column=2, value=rate)
                wb.save(DATA_PATH)
                return
        ws.append([key, rate])
        wb.save(DATA_PATH)


def api_fxrates():
    with XLSX_LOCK:
        wb = _load_wb()
        rows = _rows(wb['config'])
    rates = {}
    for r in rows:
        key = str(r.get('key') or '')
        if not key.startswith('fxrate_'):
            continue
        parts = key.split('_')
        if len(parts) != 3:
            continue
        try:
            rates[f'{parts[1]}_{parts[2]}'] = float(r.get('value') or 0)
        except (TypeError, ValueError):
            continue
    return jsonify(rates)


def modern_transfer():
    data = request.json or {}
    source_id = str(data.get('source') or '')
    dest_id = str(data.get('destination') or '')
    if not source_id or not dest_id or source_id == dest_id:
        return jsonify({'ok': False, 'error': 'source and destination must differ'}), 400
    source = get_account(source_id)
    destination = get_account(dest_id)
    if not source or not destination:
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    try:
        amount_sent = round_currency(source['currency'], abs(float(data.get('amount_sent') or 0)))
        amount_received = round_currency(destination['currency'], abs(float(data.get('amount_received') or 0)))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'invalid amount'}), 400
    if amount_sent <= 0 or amount_received <= 0:
        return jsonify({'ok': False, 'error': 'amounts must be greater than zero'}), 400

    date = str(data.get('date') or '').strip() or datetime.now().strftime('%Y-%m-%d')
    note = str(data.get('description') or '').strip()
    transfer_id = 'tx_' + secrets.token_hex(4)
    out_desc = note or f'Transfer to {destination["bank"]}'
    in_desc = note or f'Transfer from {source["bank"]}'

    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        sent_id = _next_id(ws)
        ws.append([sent_id, date, out_desc, amount_sent, 'Transfer', 'transfer', source_id, transfer_id, dest_id, 'out'])
        recv_id = _next_id(ws)
        ws.append([recv_id, date, in_desc, amount_received, 'Transfer', 'transfer', dest_id, transfer_id, source_id, 'in'])
        wb.save(DATA_PATH)

    set_balance(source_id, source['balance'] - amount_sent)
    set_balance(dest_id, destination['balance'] + amount_received)

    if source['currency'] != destination['currency'] and amount_sent > 0:
        _save_fxrate(source['currency'], destination['currency'], amount_received / amount_sent)
        _save_fxrate(destination['currency'], source['currency'], amount_sent / amount_received)

    return jsonify({'ok': True, 'transfer_id': transfer_id})


# ── RECEIPT SCANNING (AI vision) ────────────────────────────────────────────────
# Pipeline: image → AI vision model → structured fields. The receipt photo is sent
# to the user's chosen provider (Groq / Gemini / Claude), which reads it directly.
# A key is required; there is no on-device OCR fallback.

def _config_get(key, default=None):
    with XLSX_LOCK:
        wb = _load_wb()
        for r in _rows(wb['config']):
            if str(r.get('key')) == key:
                return r.get('value')
    return default

def _config_set(key, value):
    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['config']
        for row_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row=row_idx, column=1).value) == key:
                ws.cell(row=row_idx, column=2, value=value)
                break
        else:
            ws.append([key, value])
        wb.save(DATA_PATH)

def _ai_api_key():
    key = str(_config_get('ai_api_key') or '').strip()
    if key:
        return key
    return (str(os.environ.get('GROQ_API_KEY') or '').strip()
            or str(os.environ.get('ANTHROPIC_API_KEY') or '').strip()
            or str(os.environ.get('GEMINI_API_KEY') or '').strip())

def _ai_provider(api_key):
    """Auto-detect the AI provider from the key prefix. Groq keys start with
    'gsk_', Anthropic/Claude keys with 'sk-ant-'; anything else is treated as a
    Google Gemini key."""
    key = str(api_key or '')
    if key.startswith('gsk_'):
        return 'groq'
    if key.startswith('sk-ant-'):
        return 'claude'
    return 'gemini'

_HEIF_REGISTERED = False

def _register_heif():
    """Enable HEIC/HEIF decoding (the default iPhone photo format) when the
    optional pillow-heif package is installed. No-op if it isn't. Idempotent —
    safe to call on every scan."""
    global _HEIF_REGISTERED
    if _HEIF_REGISTERED:
        return
    try:
        import pillow_heif
        pillow_heif.register_heif_opener()
    except Exception:
        pass
    _HEIF_REGISTERED = True

def _normalize_fields(raw):
    raw = raw or {}
    amount = raw.get('amount')
    try:
        amount = abs(float(amount)) if amount is not None else None
    except (TypeError, ValueError):
        amount = None
    category = raw.get('category') if raw.get('category') in CATEGORIES else 'Others'
    currency = str(raw.get('currency') or '').strip().lower() or None
    if currency not in CURRENCIES:
        currency = None
    txn_type = str(raw.get('type') or '').strip().lower()
    txn_type = txn_type if txn_type in ('fund', 'expense') else 'expense'
    date = str(raw.get('date') or '').strip()
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        date = datetime.now().strftime('%Y-%m-%d')
    return {
        'amount': amount,
        'date': date,
        'merchant': str(raw.get('merchant') or '').strip()[:60],
        'category': category,
        'currency': currency,
        'type': txn_type,
    }

def _structure_prompt(text=None):
    intro = ("You read the attached photo of a store receipt and extract structured "
             "data. " if text is None else
             "You extract structured data from the raw OCR text of a store receipt. ")
    tail = ("" if text is None else
            "\nRaw OCR text:\n\"\"\"\n" + (text or '')[:6000] + "\n\"\"\"")
    return (
        intro +
        "Respond with ONLY a JSON object (no markdown, no prose) with these keys:\n"
        "  amount   - number, the grand total actually paid (not subtotal)\n"
        "  date     - 'YYYY-MM-DD' or null if not found\n"
        "  merchant - the store/vendor name, or '' if unknown\n"
        f"  category - exactly one of {CATEGORIES}. Choose by the kind of vendor:\n"
        "             - Supermarket: grocery/supermarket/convenience stores selling "
        "packaged goods, where the receipt lists many individual products "
        "(e.g. Walmart, Carrefour, Costco, Lidl, corner shop).\n"
        "             - Food: places that serve prepared meals/drinks — restaurants, "
        "cafes, bars, fast food, bakeries, delivery. Tip/cover/table lines are strong hints.\n"
        "             - Transport: fuel/gas stations, ride-hailing, taxis, parking, "
        "tolls, public transit, flights.\n"
        "             - Health: pharmacies, clinics, hospitals, dental, optical.\n"
        "             - Services: subscriptions, utilities, phone/internet, insurance, "
        "rent, repairs, gym, salons — anything billed as a service rather than goods.\n"
        "             - Games: video games, consoles, in-game purchases, gaming subscriptions.\n"
        "             - Others: only when none clearly fit.\n"
        "             Receipts are often from Uruguay — use local knowledge of merchants, e.g. "
        "Tienda Inglesa / Devoto / Disco / Ta-Ta / Multiahorro (Supermarket); La Pasiva / "
        "Bonjour / PedidosYa (Food); ANCAP / DUCSA / CUTCSA / STM (Transport); Farmashop / "
        "San Roque / CASMU (Health); Antel / UTE / OSE / Abitab / Redpagos (Services).\n"
        f"  currency - one of {list(CURRENCIES.keys())} (lowercase) or null if unknown\n"
        "  type     - 'expense' for a normal purchase, 'fund' for a refund/return/credit"
        + tail
    )

def _extract_json(out):
    match = re.search(r'\{.*\}', out or '', re.S)
    if not match:
        raise ValueError('no json in model response')
    return json.loads(match.group(0))

def _llm_complete(prompt, api_key, max_tokens=500, timeout=30, images=None):
    """Send a single user prompt (optionally with images) to the auto-detected
    provider and return the raw text of the response. Shared by receipt scanning and
    statement import — the callers parse the JSON themselves. `images` is a list of
    (mime_type, base64_data) tuples; when given, a vision-capable model is used.
    Raises urllib errors on transport/HTTP failures so callers can distinguish them."""
    provider = _ai_provider(api_key)
    if provider == 'groq':
        return _llm_complete_groq(prompt, api_key, max_tokens, timeout, images)
    if provider == 'claude':
        return _llm_complete_claude(prompt, api_key, max_tokens, timeout, images)
    return _llm_complete_gemini(prompt, api_key, max_tokens, timeout, images)

def _llm_complete_gemini(prompt, api_key, max_tokens, timeout, images=None):
    parts = [{'text': prompt}]
    for mime, b64 in (images or []):
        parts.append({'inline_data': {'mime_type': mime, 'data': b64}})
    body = json.dumps({
        'contents': [{'parts': parts}],
        'generationConfig': {'maxOutputTokens': max_tokens, 'responseMimeType': 'application/json'},
    }).encode('utf-8')
    url = ('https://generativelanguage.googleapis.com/v1beta/models/'
           + GEMINI_STRUCTURE_MODEL + ':generateContent?key=' + urllib.parse.quote(api_key))
    req = urllib.request.Request(
        url, data=body,
        headers={'content-type': 'application/json', 'user-agent': AI_USER_AGENT},
        method='POST')
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        payload = json.loads(resp.read().decode('utf-8'))
    candidates = payload.get('candidates') or []
    parts = (candidates[0].get('content', {}).get('parts') if candidates else None) or []
    return ''.join(p.get('text', '') for p in parts if isinstance(p, dict))

def _llm_complete_groq(prompt, api_key, max_tokens, timeout, images=None):
    if images:
        content = [{'type': 'text', 'text': prompt}]
        for mime, b64 in images:
            content.append({'type': 'image_url',
                            'image_url': {'url': 'data:%s;base64,%s' % (mime, b64)}})
    else:
        content = prompt
    body = json.dumps({
        'model': GROQ_VISION_MODEL if images else GROQ_STRUCTURE_MODEL,
        'messages': [{'role': 'user', 'content': content}],
        'max_tokens': max_tokens,
        'temperature': 0,
        'response_format': {'type': 'json_object'},
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.groq.com/openai/v1/chat/completions', data=body,
        headers={'content-type': 'application/json', 'authorization': 'Bearer ' + api_key,
                 'user-agent': AI_USER_AGENT},
        method='POST')
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        payload = json.loads(resp.read().decode('utf-8'))
    choices = payload.get('choices') or []
    return (choices[0].get('message', {}).get('content') if choices else '') or ''

def _llm_complete_claude(prompt, api_key, max_tokens, timeout, images=None):
    content = [{'type': 'text', 'text': prompt}]
    for mime, b64 in (images or []):
        content.append({'type': 'image',
                        'source': {'type': 'base64', 'media_type': mime, 'data': b64}})
    body = json.dumps({
        'model': CLAUDE_STRUCTURE_MODEL,
        'max_tokens': max_tokens,
        'messages': [{'role': 'user', 'content': content}],
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages', data=body,
        headers={'content-type': 'application/json', 'x-api-key': api_key,
                 'anthropic-version': '2023-06-01', 'user-agent': AI_USER_AGENT},
        method='POST')
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        payload = json.loads(resp.read().decode('utf-8'))
    blocks = payload.get('content') or []
    return ''.join(b.get('text', '') for b in blocks
                   if isinstance(b, dict) and b.get('type') == 'text')

def _ai_http_detail(exc):
    """Pull a human-readable message out of a provider's HTTP error body."""
    detail = ''
    try:
        err_body = json.loads(exc.read().decode('utf-8'))
        detail = (err_body.get('error') or {}).get('message') or ''
    except Exception:
        pass
    if exc.code == 429 and not detail:
        detail = 'rate/quota limit reached; wait a minute or check your AI plan'
    return detail

def _prepare_image_for_vision(image_bytes):
    """Decode (including HEIC), downscale and re-encode an image to a JPEG base64
    string for a vision model. Returns (mime, b64, error_code); error_code is
    'pil_unavailable' if Pillow is missing or 'bad_format' if the bytes won't open."""
    try:
        from PIL import Image, UnidentifiedImageError
    except Exception:
        return None, None, 'pil_unavailable'
    _register_heif()
    try:
        img = Image.open(io.BytesIO(image_bytes))
        img = img.convert('RGB')
        w, h = img.size
        scale = VISION_MAX_DIM / float(max(w, h))
        if scale < 1:
            img = img.resize((max(1, int(w * scale)), max(1, int(h * scale))))
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85)
        return 'image/jpeg', base64.b64encode(buf.getvalue()).decode('ascii'), None
    except UnidentifiedImageError:
        return None, None, 'bad_format'
    except Exception:
        return None, None, 'bad_format'

def _llm_structure(text, api_key):
    return _extract_json(_llm_complete(_structure_prompt(text), api_key, max_tokens=500))

def _llm_structure_vision(mime, b64, api_key):
    return _extract_json(_llm_complete(_structure_prompt(None), api_key,
                                       max_tokens=500, images=[(mime, b64)]))

def _suggest_account(currency):
    accounts = get_accounts(include_archived=False)
    if currency:
        for acc in accounts:
            if acc.get('currency') == currency:
                return acc['id']
    return accounts[0]['id'] if accounts else ''

def receipt_scan():
    file = request.files.get('image')
    if file is None:
        return jsonify({'ok': False, 'error': 'no image uploaded'}), 400
    image_bytes = file.read()
    if not image_bytes:
        return jsonify({'ok': False, 'error': 'empty image'}), 400
    if len(image_bytes) > 12 * 1024 * 1024:
        return jsonify({'ok': False, 'error': 'image too large (max 12 MB)'}), 400

    api_key = _ai_api_key()
    if not api_key:
        return jsonify({'ok': False, 'code': 'no_ai_key',
                        'error': 'Receipt scanning needs an AI key. Add one under '
                                 'Settings → AI, then try again.'}), 400

    # The receipt image goes straight to the AI vision model, which reads the photo
    # directly. No OCR: it is sharper on faded/creased/angled receipts and keeps the
    # app (and installer) light.
    mime, b64, perr = _prepare_image_for_vision(image_bytes)
    if perr == 'pil_unavailable':
        return jsonify({'ok': False, 'code': perr,
                        'error': 'Image support is not installed. Run: pip install pillow (see README).'}), 503
    if perr == 'bad_format':
        return jsonify({'ok': False, 'code': perr,
                        'error': 'This image format isn’t supported. Please use a JPG, PNG, WEBP or HEIC photo.'}), 415

    try:
        fields = _normalize_fields(_llm_structure_vision(mime, b64, api_key))
    except urllib.error.HTTPError as exc:
        detail = _ai_http_detail(exc)
        return jsonify({'ok': False, 'code': 'ai_error',
                        'error': ('AI request failed (HTTP %s).' % exc.code
                                  + (' (%s)' % detail if detail else ''))}), 502
    except urllib.error.URLError:
        return jsonify({'ok': False, 'code': 'ai_unreachable',
                        'error': 'Could not reach the AI service. Check your connection and try again.'}), 502
    except Exception:
        return jsonify({'ok': False, 'code': 'ai_failed',
                        'error': 'The AI could not read this receipt. Try a clearer, straight-on photo.'}), 502

    fields['suggested_account'] = _suggest_account(fields.get('currency'))
    return jsonify({'ok': True, 'method': 'vision', 'warning': None,
                    'fields': fields, 'raw_text': ''})

def receipt_config_get():
    saved = str(_config_get('ai_api_key') or '').strip()
    env_key = (str(os.environ.get('GROQ_API_KEY') or '').strip()
               or str(os.environ.get('ANTHROPIC_API_KEY') or '').strip()
               or str(os.environ.get('GEMINI_API_KEY') or '').strip())
    active = saved or (env_key if not saved else '')
    provider = _ai_provider(active) if active else None
    model = {'groq': GROQ_STRUCTURE_MODEL, 'claude': CLAUDE_STRUCTURE_MODEL,
             'gemini': GEMINI_STRUCTURE_MODEL}.get(provider, GEMINI_STRUCTURE_MODEL)
    return jsonify({
        'ok': True,
        'has_key': bool(saved),
        'env_key': bool(env_key) and not saved,
        'provider': provider,
        'provider_name': {'groq': 'Groq', 'claude': 'Claude', 'gemini': 'Google Gemini'}.get(provider, ''),
        'model': model,
        'key_hint': ('…' + saved[-4:]) if len(saved) >= 4 else '',
    })

def _verify_ai_key(api_key):
    """Cheaply check that a key authenticates with its provider. Hits the
    provider's models-list endpoint (GET, costs no tokens). Returns (True, '')
    on success or (False, 'reason') on failure."""
    provider = _ai_provider(api_key)
    try:
        if provider == 'groq':
            req = urllib.request.Request(
                'https://api.groq.com/openai/v1/models',
                headers={'authorization': 'Bearer ' + api_key, 'user-agent': AI_USER_AGENT})
        elif provider == 'claude':
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/models',
                headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01',
                         'user-agent': AI_USER_AGENT})
        else:
            req = urllib.request.Request(
                'https://generativelanguage.googleapis.com/v1beta/models?key='
                + urllib.parse.quote(api_key),
                headers={'user-agent': AI_USER_AGENT})
        with urllib.request.urlopen(req, timeout=15) as resp:
            resp.read()
        return True, ''
    except urllib.error.HTTPError as exc:
        detail = ''
        try:
            err = json.loads(exc.read().decode('utf-8')).get('error')
            detail = err.get('message') if isinstance(err, dict) else (err or '')
        except Exception:
            pass
        return False, (detail or ('HTTP %s' % exc.code))
    except urllib.error.URLError:
        return False, 'could not reach the provider'
    except Exception:
        return False, 'verification failed'

def receipt_config_set():
    data = request.json or {}
    key = str(data.get('api_key') or '').strip()
    provider = _ai_provider(key) if key else None
    prov_name = {'groq': 'Groq', 'claude': 'Claude', 'gemini': 'Google Gemini'}.get(provider, 'AI')
    if key:
        ok, reason = _verify_ai_key(key)
        if not ok:
            return jsonify({'ok': False, 'has_key': False, 'verified': False,
                            'provider': provider,
                            'error': prov_name + ' rejected the key: ' + reason}), 400
    _config_set('ai_api_key', key)
    return jsonify({'ok': True, 'has_key': bool(key), 'verified': bool(key),
                    'provider': provider, 'provider_name': prov_name if key else ''})


# ── BANK STATEMENT IMPORT (PDF → many transactions) ──────────────────────────────
# Pipeline: PDF → pypdf text extraction (local) → AI structuring into an array of
# movements → multi-row review → batch create. Requires an AI key: bank statement
# layouts vary too much between banks for a reliable generic parser, so there is no
# regex fallback. The user picks the target account up front (one statement = one
# account). The PDF text never leaves the machine except the call to the AI provider.

STATEMENT_MAX_PDF_BYTES = 20 * 1024 * 1024
STATEMENT_TEXT_LIMIT = 24000   # chars of extracted PDF text sent to the model
STATEMENT_MAX_TOKENS = 4000

def _pdf_text(pdf_bytes):
    """Return (text, truncated, error_code). error_code is None on success.
    'pdf_unavailable' if pypdf isn't installed, 'bad_pdf' if it won't open,
    'no_text' if there's no extractable text (likely a scanned/image PDF)."""
    try:
        from pypdf import PdfReader
    except Exception:
        return None, False, 'pdf_unavailable'
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        parts = []
        for page in reader.pages:
            try:
                parts.append(page.extract_text() or '')
            except Exception:
                continue
        text = '\n'.join(parts).strip()
    except Exception:
        return None, False, 'bad_pdf'
    if not text:
        return None, False, 'no_text'
    truncated = len(text) > STATEMENT_TEXT_LIMIT
    return text[:STATEMENT_TEXT_LIMIT], truncated, None

def _statement_prompt(text=None):
    source = ("the attached page images of a bank or credit-card statement"
              if text is None else
              "the raw text of a bank or credit-card statement")
    tail = ("" if text is None else
            "\n\nRaw statement text:\n\"\"\"\n" + (text or '') + "\n\"\"\"")
    return (
        "You extract every transaction from " + source + ". "
        "Respond with ONLY a JSON object (no markdown, no prose) of the "
        'form {"transactions": [ ... ]}, where each array item has these keys:\n'
        "  date        - 'YYYY-MM-DD'. Infer the year from the statement period when "
        "a row only shows day/month.\n"
        "  description - the merchant or description of the movement, trimmed.\n"
        "  amount      - a positive number (the movement amount, never negative).\n"
        "  type        - 'expense' for money leaving the account (debits, charges, "
        "purchases, withdrawals), 'fund' for money coming in (credits, deposits, "
        "refunds, incoming transfers). IMPORTANT: when the statement has a running "
        "balance/saldo column, use it as the source of truth. Read the rows in date "
        "order and compare each row's balance to the previous row's: if the balance "
        "went UP the movement is a 'fund' (credit), if it went DOWN it is an "
        "'expense' (debit). Trust the balance over the wording: merchant names like "
        "supermarkets or restaurants can appear on credit lines too (for example tax "
        "refunds), so a row mentioning a store is not automatically an expense.\n"
        "  iva_refund  - true ONLY for Uruguayan IVA-refund credits: small 'fund' "
        "lines such as 'REDIVA', 'Reintegro de IVA' or 'Devolucion de IVA' that the "
        "bank gives back for paying by card. false for every other movement.\n"
        f"  category    - exactly one of {CATEGORIES}, chosen by the kind of vendor:\n"
        "                Supermarket (grocery/convenience), Food (restaurants, cafes, "
        "bars, fast food, delivery), Transport (fuel, ride-hailing, taxis, parking, "
        "tolls, transit, flights), Health (pharmacies, clinics, dental, optical), "
        "Services (subscriptions, utilities, phone/internet, insurance, rent, repairs, "
        "gym, bank fees), Games (video games, consoles, gaming subscriptions), "
        "'Hanging out' (leisure, entertainment, shopping for fun), Others (only when "
        "none clearly fit).\n"
        "                Statements are often from Uruguay — use local knowledge: "
        "Tienda Inglesa / Devoto / Disco / Ta-Ta / Multiahorro (Supermarket); La Pasiva "
        "/ Bonjour / PedidosYa (Food); ANCAP / DUCSA / CUTCSA / STM (Transport); "
        "Farmashop / San Roque / CASMU (Health); Antel / UTE / OSE / Abitab / Redpagos "
        "(Services).\n"
        "Skip every non-transaction line: opening/closing balances, totals, subtotals, "
        "interest summaries, and any row without a real movement amount. Keep the "
        "order they appear. If there are no transactions, return "
        '{"transactions": []}.'
        + tail
    )

def _extract_json_array(out):
    """Parse the model output into a list of transaction dicts. Accepts either a
    bare JSON array or an object wrapping the array under a transactions-like key."""
    decoder = json.JSONDecoder()
    s = out or ''
    for i, ch in enumerate(s):
        if ch not in '[{':
            continue
        try:
            data, _ = decoder.raw_decode(s[i:])
        except ValueError:
            continue
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for key in ('transactions', 'items', 'movements', 'data'):
                if isinstance(data.get(key), list):
                    return data[key]
        # parsed a JSON value but not the array we want — keep scanning
    raise ValueError('no transactions array in model response')

def _llm_statement(text, api_key):
    out = _llm_complete(_statement_prompt(text), api_key,
                        max_tokens=STATEMENT_MAX_TOKENS, timeout=60)
    return _extract_json_array(out)

def _pdf_page_images(pdf_bytes):
    """For a scanned/image PDF (no extractable text), pull the embedded page images
    and return them as a list of (mime, base64) JPEGs for a vision model, plus a
    `truncated` flag if there were more pages than VISION_MAX_PAGES. Returns an empty
    list when pypdf/Pillow are missing or the pages carry no usable image."""
    try:
        from pypdf import PdfReader
    except Exception:
        return [], False
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return [], False
    out, truncated = [], False
    for page in reader.pages:
        if len(out) >= VISION_MAX_PAGES:
            truncated = True
            break
        try:
            page_imgs = list(page.images)
        except Exception:
            page_imgs = []
        if not page_imgs:
            continue
        biggest = max(page_imgs, key=lambda im: len(getattr(im, 'data', b'') or b''))
        mime, b64, perr = _prepare_image_for_vision(biggest.data)
        if perr is None:
            out.append((mime, b64))
    return out, truncated

def _llm_statement_vision(images, api_key):
    out = _llm_complete(_statement_prompt(None), api_key,
                        max_tokens=STATEMENT_MAX_TOKENS, timeout=120, images=images)
    return _extract_json_array(out)

def _normalize_statement_item(raw, currency):
    """Coerce one model item into a clean transaction dict, or None to drop it.
    `currency` is the target account's lowercase currency, used for rounding."""
    if not isinstance(raw, dict):
        return None
    try:
        amount = round_currency(currency, abs(float(raw.get('amount'))))
    except (TypeError, ValueError):
        return None
    if amount <= 0:
        return None
    date = str(raw.get('date') or '').strip()
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        return None
    txn_type = str(raw.get('type') or '').strip().lower()
    txn_type = txn_type if txn_type in ('fund', 'expense') else 'expense'
    category = str(raw.get('category') or '').strip()
    if category not in CATEGORIES:
        category = 'Others'
    description = str(raw.get('description') or '').strip()[:60]
    iva_refund = bool(raw.get('iva_refund')) and txn_type == 'fund'
    return {'date': date, 'description': description, 'amount': amount,
            'type': txn_type, 'category': category, 'iva_refund': iva_refund}

def _consolidate_iva_refunds(items, currency):
    """Collapse all Uruguayan IVA-refund credits (REDIVA / Reintegro de IVA) into a
    single 'fund' movement, so the statement shows one income row instead of many
    tiny tax refunds. The merged row keeps the position of the first refund, sums
    the amounts (in code, so the total is exact), and is dated by the last refund.
    Non-refund items are left untouched and keep their order."""
    refunds = [it for it in items if it.get('iva_refund')]
    if len(refunds) < 2:
        return items
    merged = {
        'date': max(it['date'] for it in refunds),
        'description': 'Reintegro de IVA',
        'amount': round_currency(currency, sum(it['amount'] for it in refunds)),
        'type': 'fund',
        'category': 'Others',
        'iva_refund': True,
    }
    out, inserted = [], False
    for it in items:
        if it.get('iva_refund'):
            if not inserted:
                out.append(merged)
                inserted = True
            continue
        out.append(it)
    return out

def _flag_duplicates(items, account_id, currency):
    """Mark each item that already exists as a transaction in this account (same
    date, type and amount) with duplicate=True and include=False. Description is
    intentionally ignored — the bank's wording differs from manual/scanned entries."""
    with XLSX_LOCK:
        wb = _load_wb()
        rows = _rows(wb['transactions'])
    existing = set()
    for r in rows:
        if str(r.get('account') or '') != account_id:
            continue
        try:
            amt = round_currency(currency, abs(float(r.get('amount') or 0)))
        except (TypeError, ValueError):
            continue
        existing.add((str(r.get('date') or ''), str(r.get('type') or ''), amt))
    for it in items:
        it['duplicate'] = (it['date'], it['type'], it['amount']) in existing
        it['include'] = not it['duplicate']
    return items

def statement_scan():
    file = request.files.get('pdf')
    if file is None:
        return jsonify({'ok': False, 'error': 'no PDF uploaded'}), 400
    account = get_account(str(request.form.get('account') or '').strip())
    if not account:
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    api_key = _ai_api_key()
    if not api_key:
        return jsonify({'ok': False, 'code': 'no_ai_key',
                        'error': 'Statement import needs an AI key. Add one under '
                                 'Settings → Receipt Scanning, then try again.'}), 400
    pdf_bytes = file.read()
    if not pdf_bytes:
        return jsonify({'ok': False, 'error': 'empty PDF'}), 400
    if len(pdf_bytes) > STATEMENT_MAX_PDF_BYTES:
        return jsonify({'ok': False, 'error': 'PDF too large (max 20 MB)'}), 400

    text, truncated, err = _pdf_text(pdf_bytes)
    if err == 'pdf_unavailable':
        return jsonify({'ok': False, 'code': err,
                        'error': 'PDF support is not installed. Run: pip install pypdf (see README).'}), 503
    if err == 'bad_pdf':
        return jsonify({'ok': False, 'code': err,
                        'error': 'Could not read this PDF. Make sure it is a valid, unprotected PDF.'}), 422
    if err and err != 'no_text':
        return jsonify({'ok': False, 'code': err, 'error': 'Could not read the PDF.'}), 500

    # Text PDFs (the common case) read the extracted text, which is exact and cheap.
    # A scanned/image PDF has no text, so fall back to sending the page images to a
    # vision model.
    images = None
    if err == 'no_text':
        images, img_truncated = _pdf_page_images(pdf_bytes)
        if not images:
            return jsonify({'ok': False, 'code': 'no_text',
                            'error': 'This PDF has no selectable text and no readable page images. '
                                     'Export the statement as a text-based PDF and try again.'}), 422
        truncated = truncated or img_truncated

    try:
        if images:
            raw_items = _llm_statement_vision(images, api_key)
            method = 'vision'
        else:
            raw_items = _llm_statement(text, api_key)
            method = 'text'
    except urllib.error.HTTPError as exc:
        detail = _ai_http_detail(exc)
        return jsonify({'ok': False, 'code': 'ai_error',
                        'error': ('AI request failed (HTTP %s).' % exc.code
                                  + (' (%s)' % detail if detail else ''))}), 502
    except urllib.error.URLError:
        return jsonify({'ok': False, 'code': 'ai_unreachable',
                        'error': 'Could not reach the AI service. Check your connection and try again.'}), 502
    except Exception:
        return jsonify({'ok': False, 'code': 'ai_failed',
                        'error': 'The AI could not extract transactions from this statement.'}), 502

    currency = account['currency']
    items = [n for n in (_normalize_statement_item(r, currency) for r in raw_items) if n]
    items = _consolidate_iva_refunds(items, currency)
    _flag_duplicates(items, account['id'], currency)
    return jsonify({'ok': True, 'count': len(items), 'truncated': truncated,
                    'method': method, 'account': account['id'], 'currency': currency,
                    'transactions': items})

def statement_import():
    data = request.json or {}
    account = get_account(str(data.get('account') or '').strip())
    if not account:
        return jsonify({'ok': False, 'error': 'unknown account'}), 400
    raw_items = data.get('transactions')
    if not isinstance(raw_items, list) or not raw_items:
        return jsonify({'ok': False, 'error': 'no transactions to import'}), 400

    currency = account['currency']
    items = [n for n in (_normalize_statement_item(r, currency) for r in raw_items) if n]
    if not items:
        return jsonify({'ok': False, 'error': 'no valid transactions to import'}), 400

    # Adjust the balance first: set_balance acquires XLSX_LOCK internally, so it must
    # run outside the append block below to avoid the non-reentrant deadlock — same
    # ordering as _add_txn (balance before the row write).
    delta = sum(it['amount'] if it['type'] == 'fund' else -it['amount'] for it in items)
    new_balance = set_balance(account['id'], get_balances()[account['id']] + delta)

    with XLSX_LOCK:
        wb = _load_wb()
        ws = wb['transactions']
        for it in items:
            ws.append([
                _next_id(ws), it['date'], it['description'], it['amount'],
                it['category'], it['type'], account['id'],
            ])
        wb.save(DATA_PATH)
    return jsonify({'ok': True, 'created': len(items), 'balance': new_balance})


app.add_url_rule('/api/transfer', 'modern_transfer', modern_transfer, methods=['POST'])
app.add_url_rule('/api/receipt/scan', 'receipt_scan', receipt_scan, methods=['POST'])
app.add_url_rule('/api/receipt/config', 'receipt_config_get', receipt_config_get, methods=['GET'])
app.add_url_rule('/api/receipt/config', 'receipt_config_set', receipt_config_set, methods=['POST'])
app.add_url_rule('/api/statement/scan', 'statement_scan', statement_scan, methods=['POST'])
app.add_url_rule('/api/statement/import', 'statement_import', statement_import, methods=['POST'])
app.add_url_rule('/api/fxrates', 'api_fxrates', api_fxrates, methods=['GET'])
app.add_url_rule('/api/transactions/<int:tid>', 'delete_txn', modern_delete_txn, methods=['DELETE'])
app.add_url_rule('/api/transactions/<int:tid>', 'edit_txn', modern_edit_txn, methods=['PUT'])
app.add_url_rule('/api/balance', 'api_set_balance', modern_set_balance, methods=['POST'])
app.add_url_rule('/api/export', 'api_export', modern_export, methods=['GET'])
app.add_url_rule('/api/import', 'api_import', modern_import, methods=['POST'])
app.add_url_rule('/api/clear', 'api_clear', modern_clear, methods=['POST'])
app.add_url_rule('/api/fixed', 'api_fixed', modern_fixed, methods=['GET'])
app.add_url_rule('/api/fixed', 'create_fixed', modern_create_fixed, methods=['POST'])
app.add_url_rule('/api/fixed/<int:fid>', 'edit_fixed', modern_edit_fixed, methods=['PUT'])
app.add_url_rule('/api/accounts', 'modern_accounts', modern_accounts, methods=['GET'])
app.add_url_rule('/api/accounts', 'modern_create_account', modern_create_account, methods=['POST'])
app.add_url_rule('/api/accounts/<account_id>', 'modern_edit_account', modern_edit_account, methods=['PUT'])
app.add_url_rule('/api/accounts/<account_id>', 'modern_delete_account', modern_delete_account, methods=['DELETE'])
app.add_url_rule('/api/accounts/<account_id>/permanent', 'modern_permanent_delete_account', modern_permanent_delete_account, methods=['DELETE'])


def _parse_semver(tag):
    s = (tag or '').lstrip('vV').strip()
    parts = s.split('-', 1)[0].split('.')
    out = []
    for p in parts:
        try:
            out.append(int(p))
        except ValueError:
            out.append(0)
    while len(out) < 3:
        out.append(0)
    return tuple(out[:3])


@app.route('/api/version/check')
def api_version_check():
    url = f'https://api.github.com/repos/{GITHUB_REPO}/releases/latest'
    req = urllib.request.Request(url, headers={
        'Accept': 'application/vnd.github+json',
        'User-Agent': 'ClariFi-Updater',
    })
    try:
        with urllib.request.urlopen(req, timeout=6) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return jsonify({
                'ok': True, 'current': APP_VERSION, 'latest': None,
                'update_available': False, 'message': 'No releases published yet',
            })
        return jsonify({'ok': False, 'current': APP_VERSION, 'error': f'github error {e.code}'}), 502
    except (urllib.error.URLError, TimeoutError, OSError):
        return jsonify({'ok': False, 'current': APP_VERSION, 'error': 'could not reach update server'}), 502
    except (ValueError, KeyError):
        return jsonify({'ok': False, 'current': APP_VERSION, 'error': 'invalid response from update server'}), 502

    latest_tag = data.get('tag_name') or ''
    html_url = data.get('html_url') or f'https://github.com/{GITHUB_REPO}/releases'
    published = data.get('published_at')
    notes = data.get('body') or ''

    assets = data.get('assets') or []
    installer = None
    for a in assets:
        name = (a.get('name') or '').lower()
        if name.endswith('.exe') or name.endswith('.msi'):
            installer = a.get('browser_download_url')
            break

    update_available = _parse_semver(latest_tag) > _parse_semver(APP_VERSION)

    return jsonify({
        'ok': True,
        'current': APP_VERSION,
        'latest': latest_tag,
        'update_available': update_available,
        'release_url': html_url,
        'installer_url': installer,
        'published_at': published,
        'notes': notes,
        'repo': GITHUB_REPO,
    })

_DOWNLOAD_STATE = {
    'status': 'idle',     # idle | downloading | done | error
    'downloaded': 0,
    'total': 0,
    'path': None,
    'error': None,
}
_DOWNLOAD_LOCK = Lock()


def _download_worker(url, dest):
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'ClariFi-Updater'})
        with urllib.request.urlopen(req, timeout=120) as resp, open(dest, 'wb') as f:
            try:
                total = int(resp.headers.get('Content-Length') or 0)
            except (TypeError, ValueError):
                total = 0
            with _DOWNLOAD_LOCK:
                _DOWNLOAD_STATE['total'] = total
            while True:
                chunk = resp.read(65536)
                if not chunk:
                    break
                f.write(chunk)
                with _DOWNLOAD_LOCK:
                    _DOWNLOAD_STATE['downloaded'] += len(chunk)
        with _DOWNLOAD_LOCK:
            _DOWNLOAD_STATE['status'] = 'done'
            _DOWNLOAD_STATE['path'] = dest
    except (urllib.error.URLError, TimeoutError, OSError) as e:
        with _DOWNLOAD_LOCK:
            _DOWNLOAD_STATE['status'] = 'error'
            _DOWNLOAD_STATE['error'] = f'download failed: {e}'


@app.route('/api/version/download', methods=['POST'])
def api_version_download():
    data = request.json or {}
    url = data.get('url')
    if not url or not isinstance(url, str) or not url.startswith('https://'):
        return jsonify({'ok': False, 'error': 'invalid installer url'}), 400
    if not (url.endswith('.exe') or url.endswith('.msi')):
        return jsonify({'ok': False, 'error': 'unexpected installer extension'}), 400

    import tempfile, threading
    with _DOWNLOAD_LOCK:
        if _DOWNLOAD_STATE['status'] == 'downloading':
            return jsonify({'ok': False, 'error': 'a download is already in progress'}), 400
        _DOWNLOAD_STATE.update({
            'status': 'downloading',
            'downloaded': 0,
            'total': 0,
            'path': None,
            'error': None,
        })

    dest = os.path.join(tempfile.gettempdir(), os.path.basename(url))
    threading.Thread(target=_download_worker, args=(url, dest), daemon=True).start()
    return jsonify({'ok': True})


@app.route('/api/version/download/progress')
def api_version_download_progress():
    with _DOWNLOAD_LOCK:
        return jsonify(dict(_DOWNLOAD_STATE))


@app.route('/api/version/install', methods=['POST'])
def api_version_install():
    data = request.json or {}
    path = data.get('path') or ''
    if not path or not os.path.isfile(path):
        return jsonify({'ok': False, 'error': 'installer file not found'}), 400
    if not (path.lower().endswith('.exe') or path.lower().endswith('.msi')):
        return jsonify({'ok': False, 'error': 'unexpected installer extension'}), 400

    import subprocess, threading
    flags = 0
    if os.name == 'nt':
        flags = getattr(subprocess, 'DETACHED_PROCESS', 0) | getattr(subprocess, 'CREATE_NEW_PROCESS_GROUP', 0)
    try:
        subprocess.Popen(
            [path, '/SILENT', '/SUPPRESSMSGBOXES'],
            creationflags=flags,
            close_fds=True,
        )
    except OSError as e:
        return jsonify({'ok': False, 'error': f'could not launch installer: {e}'}), 500

    # Give the response time to flush, then kill ourselves so the installer
    # can replace our files. The installer's [Run] entry relaunches the app.
    threading.Timer(1.2, lambda: os._exit(0)).start()
    return jsonify({'ok': True})


if __name__ == '__main__':
    init_data()
    port = int(os.environ.get('PORT', 5000))
    print(
        '\n'
        f'  ClariFi Desktop -> http://localhost:{port}\n'
    )
    app.run(host='127.0.0.1', port=port, debug=False)
